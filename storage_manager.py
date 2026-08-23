"""
消防及義消子女獎學金 AI 智慧審核系統 - 自動儲存與檔案持久化管理模組
負責將同仁上傳照片存入 ./uploads/，並將審核資料自動追加 (append) 寫入 ./data/獎學金總表.xlsx 與 ./data/records.json
"""

import os
import io
import json
import zipfile
from datetime import datetime
from typing import List, Dict, Any, Tuple
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
DATA_DIR = os.path.join(BASE_DIR, "data")
EXCEL_FILE = os.path.join(DATA_DIR, "獎學金總表.xlsx")
JSON_FILE = os.path.join(DATA_DIR, "records.json")

def ensure_directories():
    """確保 uploads 與 data 資料夾存在"""
    os.makedirs(UPLOADS_DIR, exist_ok=True)
    os.makedirs(DATA_DIR, exist_ok=True)

def save_case_to_storage(case_dict: Dict[str, Any]) -> Tuple[bool, str]:
    """
    1. 將照片儲存至 ./uploads/{case_id}_{index}.jpg
    2. 將紀錄更新/追加至 ./data/records.json
    3. 自動追加 (append) 至 ./data/獎學金總表.xlsx
    """
    ensure_directories()
    case_id = case_dict.get("id", f"TTFD-{datetime.now().strftime('%Y%m%d%H%M%S')}")
    
    # 1. 儲存圖片至 ./uploads/
    saved_paths = []
    images = case_dict.get("images", [])
    for idx, img in enumerate(images, 1):
        filename = f"{case_id}_page{idx}.jpg"
        full_path = os.path.join(UPLOADS_DIR, filename)
        try:
            if isinstance(img, Image.Image):
                # 轉為 RGB 存檔
                rgb_img = img.convert("RGB") if img.mode in ("RGBA", "P") else img
                rgb_img.save(full_path, "JPEG", quality=92)
                saved_paths.append(filename)
        except Exception as e:
            print(f"Error saving image {filename}: {e}")
            
    case_dict["image_paths"] = saved_paths
    
    # 2. 讀取並追加至 records.json
    all_records = []
    if os.path.exists(JSON_FILE):
        try:
            with open(JSON_FILE, "r", encoding="utf-8") as f:
                all_records = json.load(f)
        except Exception:
            all_records = []
            
    # 序列化處理 (去除無法轉 json 的 PIL 物件)
    json_record = {k: v for k, v in case_dict.items() if k not in ("images",)}
    
    # 檢查是否已存在同案號 (若存在則更新，否則 append)
    existing_idx = next((i for i, r in enumerate(all_records) if r.get("id") == case_id), None)
    if existing_idx is not None:
        all_records[existing_idx] = json_record
    else:
        all_records.append(json_record)
        
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(all_records, f, ensure_ascii=False, indent=2)
        
    # 3. 追加寫入 ./data/獎學金總表.xlsx
    try:
        append_case_to_excel(case_dict, len(all_records))
    except Exception as e:
        print(f"Error appending to Excel: {e}")
        
    return True, f"已成功將 {len(saved_paths)} 張照片存入 ./uploads/，並將數據追加至 ./data/獎學金總表.xlsx！"

def append_case_to_excel(case_dict: Dict[str, Any], seq_num: int):
    """將單筆案件追加寫入 Excel 總表"""
    from excel_exporter import export_scholarship_excel
    
    ensure_directories()
    
    # 若 Excel 還不存在，直接完整生成
    if not os.path.exists(EXCEL_FILE):
        excel_bytes = export_scholarship_excel([case_dict])
        with open(EXCEL_FILE, "wb") as f:
            f.write(excel_bytes.getvalue())
        return
        
    # 若已存在，使用 openpyxl 載入並追加列
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    except Exception:
        # 若檔案損毀則重新生成
        excel_bytes = export_scholarship_excel([case_dict])
        with open(EXCEL_FILE, "wb") as f:
            f.write(excel_bytes.getvalue())
        return
        
    font_data = Font(name="微軟正黑體", size=10, color="000000")
    thin_border_side = Side(style='thin', color='D3D3D3')
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_eligible = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    fill_pending = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    fill_ineligible = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    font_status_eligible = Font(name="微軟正黑體", size=10, bold=True, color="155724")
    font_status_pending = Font(name="微軟正黑體", size=10, bold=True, color="856404")
    font_status_ineligible = Font(name="微軟正黑體", size=10, bold=True, color="721C24")
    
    att = case_dict.get("attachments", {})
    att_keys = [
        ("application_form", "申請表"),
        ("student_id_or_enrollment", "在學證明"),
        ("transcript", "成績單"),
        ("household_registration", "戶籍資料"),
        ("service_certificate", "服務證明")
    ]
    present_count = sum(1 for k, _ in att_keys if att.get(k, False))
    if present_count == 5:
        att_text = "齊全 (5/5)"
    else:
        missing_names = [name for k, name in att_keys if not att.get(k, False)]
        att_text = f"缺: {', '.join(missing_names)} ({present_count}/5)"
        
    status = case_dict.get("review_status", "待審核")
    gpa = case_dict.get("semester_gpa")
    gpa_val = f"{gpa:.2f}" if gpa is not None else "無"
    
    # 計算追加列號 (表頭在第 6 列，第 7 列起為資料)
    curr_row = max(ws.max_row + 1, 7)
    
    row_values = [
        seq_num,
        case_dict.get("id", ""),
        case_dict.get("unit_level1", "未指定"),
        case_dict.get("unit_level2", "未指定"),
        case_dict.get("applicant_name", ""),
        case_dict.get("applicant_id", ""),
        case_dict.get("child_name", ""),
        case_dict.get("category", ""),
        gpa_val,
        case_dict.get("conduct", ""),
        att_text,
        status,
        case_dict.get("review_reason", case_dict.get("notes", "")),
        ""
    ]
    
    for col_idx, val in enumerate(row_values, 1):
        cell = ws.cell(row=curr_row, column=col_idx, value=val)
        cell.font = font_data
        cell.border = border_cell
        
        if seq_num % 2 == 0:
            cell.fill = fill_zebra
            
        if col_idx in (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14):
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
        if col_idx == 12:
            if status == "符合資格":
                cell.font = font_status_eligible
                cell.fill = fill_eligible
            elif status == "待補件":
                cell.font = font_status_pending
                cell.fill = fill_pending
            elif status == "不符資格":
                cell.font = font_status_ineligible
                cell.fill = fill_ineligible
                
    ws.row_dimensions[curr_row].height = 24
    
    # 更新 Row 2 匯出時間與統計摘要列 (Row 3, 4)
    ws["A2"] = f"匯出機關：臺東縣消防局  |  最後更新：{datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}  |  審核門檻：學期總平均 >= 80.0 且 5 項附件齊全"
    
    wb.save(EXCEL_FILE)

def load_stored_cases() -> List[Dict[str, Any]]:
    """從 ./data/records.json 與 ./uploads/ 載入所有歷史儲存案件"""
    ensure_directories()
    if not os.path.exists(JSON_FILE):
        return []
        
    try:
        with open(JSON_FILE, "r", encoding="utf-8") as f:
            records = json.load(f)
            
        for r in records:
            images = []
            image_labels = []
            for path_fn in r.get("image_paths", []):
                full_p = os.path.join(UPLOADS_DIR, path_fn)
                if os.path.exists(full_p):
                    try:
                        img = Image.open(full_p)
                        images.append(img)
                        image_labels.append(path_fn)
                    except Exception:
                        pass
            r["images"] = images
            r["image_labels"] = image_labels
            
        return records
    except Exception as e:
        print(f"Error loading stored cases: {e}")
        return []

def package_uploads_zip() -> io.BytesIO:
    """將整個 ./uploads/ 目錄打包為 zip 檔案供承辦人下載"""
    ensure_directories()
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(UPLOADS_DIR):
            for file in files:
                full_path = os.path.join(root, file)
                rel_path = os.path.relpath(full_path, UPLOADS_DIR)
                z.write(full_path, rel_path)
    zip_buf.seek(0)
    return zip_buf
