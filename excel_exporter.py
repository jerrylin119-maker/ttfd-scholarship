"""
消防及義消子女獎學金 AI 智慧審核系統 - Excel 格式化匯出模組 (含案件編號與兩階層組織單位)
"""

import io
from datetime import datetime
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def export_scholarship_excel(records: List[Dict[str, Any]]) -> io.BytesIO:
    """
    將審核紀錄匯出為專業格式化的 Excel 活頁簿 (.xlsx)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "獎學金審核清冊"
    
    ws.views.sheetView[0].showGridLines = True
    
    font_title = Font(name="微軟正黑體", size=16, bold=True, color="1F4E79")
    font_subtitle = Font(name="微軟正黑體", size=10, italic=True, color="595959")
    font_summary_label = Font(name="微軟正黑體", size=10, bold=True, color="333333")
    font_summary_val = Font(name="微軟正黑體", size=11, bold=True, color="1F4E79")
    font_header = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="微軟正黑體", size=10, color="000000")
    font_status_eligible = Font(name="微軟正黑體", size=10, bold=True, color="155724")
    font_status_pending = Font(name="微軟正黑體", size=10, bold=True, color="856404")
    font_status_ineligible = Font(name="微軟正黑體", size=10, bold=True, color="721C24")
    
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_sub_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_eligible = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    fill_pending = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    fill_ineligible = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    thin_border_side = Side(style='thin', color='D3D3D3')
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_header = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=Side(style='medium', color='1F4E79'))
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    total_count = len(records)
    eligible_count = sum(1 for r in records if r.get("review_status") == "符合資格")
    pending_count = sum(1 for r in records if r.get("review_status") == "待補件")
    ineligible_count = sum(1 for r in records if r.get("review_status") == "不符資格")
    pass_rate = f"{(eligible_count / total_count * 100):.1f}%" if total_count > 0 else "0.0%"
    
    # 1. 標題與副標題 (A1:N1, A2:N2)
    ws.merge_cells("A1:N1")
    ws["A1"] = "🚒 臺東縣消防局 消防及義消人員子女獎學金 AI 智慧審核彙總名冊"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells("A2:N2")
    now_str = datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")
    ws["A2"] = f"匯出機關：臺東縣消防局  |  匯出時間：{now_str}  |  審核門檻：學期總平均 >= 80.0 且 5 項附件齊全"
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = align_center
    ws.row_dimensions[2].height = 20
    
    # 2. 統計摘要列 (Row 3 & 4)
    summary_items = [
        ("A3", "B3", "A4", "B4", "總申請件數", f"{total_count} 件"),
        ("C3", "E3", "C4", "E4", "符合資格", f"{eligible_count} 件"),
        ("F3", "H3", "F4", "H4", "待補件", f"{pending_count} 件"),
        ("I3", "K3", "I4", "K4", "不符資格", f"{ineligible_count} 件"),
        ("L3", "N3", "L4", "N4", "核定通過率", pass_rate)
    ]
    
    for l_start, l_end, v_start, v_end, label, val in summary_items:
        ws.merge_cells(f"{l_start}:{l_end}")
        ws[l_start] = label
        ws[l_start].font = font_summary_label
        ws[l_start].alignment = align_center
        ws[l_start].fill = fill_sub_header
        
        ws.merge_cells(f"{v_start}:{v_end}")
        ws[v_start] = val
        ws[v_start].font = font_summary_val
        ws[v_start].alignment = align_center
        ws[v_start].fill = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
    
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 10
    
    # 3. 表頭 (Row 6) - 含「案件編號」
    headers = [
        "序號", "案件編號", "大隊 / 局本部", "分隊 / 科室", "申請人姓名", "身分證字號", "子女姓名", "申請組別",
        "學期總平均", "操行成績", "附件檢核狀況 (5項)", "審核結果", "審核判定說明 / 備註", "承辦人複核簽章"
    ]
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header
    
    ws.row_dimensions[6].height = 28
    
    # 4. 資料內容 (Row 7+)
    start_row = 7
    for idx, r in enumerate(records, 1):
        curr_row = start_row + idx - 1
        
        att = r.get("attachments", {})
        att_keys = [
            ("application_form", "申請表"),
            ("student_id_or_enrollment", "在學證明"),
            ("transcript", "成績單"),
            ("household_registration", "戶籍資料"),
            ("service_certificate", "服務證明")
        ]
        present_count = sum(1 for k, _ in att_keys if att.get(k, False))
        if present_count == 5:
            att_text = "齊全 (5/5)"
        else:
            missing_names = [name for k, name in att_keys if not att.get(k, False)]
            att_text = f"缺: {', '.join(missing_names)} ({present_count}/5)"
            
        status = r.get("review_status", "待審核")
        gpa = r.get("semester_gpa")
        gpa_val = f"{gpa:.2f}" if gpa is not None else "無"
        
        row_values = [
            idx,
            r.get("id", f"CASE-{idx:03d}"),
            r.get("unit_level1", "未指定"),
            r.get("unit_level2", "未指定"),
            r.get("applicant_name", ""),
            r.get("applicant_id", ""),
            r.get("child_name", ""),
            r.get("category", ""),
            gpa_val,
            r.get("conduct", ""),
            att_text,
            status,
            r.get("review_reason", r.get("notes", "")),
            ""
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=curr_row, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_cell
            
            if idx % 2 == 0:
                cell.fill = fill_zebra
                
            # 文字對齊：判定說明靠左，其餘置中
            if col_idx in (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14):
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            if col_idx == 12:  # 審核結果欄位顏色
                if status == "符合資格":
                    cell.font = font_status_eligible
                    cell.fill = fill_eligible
                elif status == "待補件":
                    cell.font = font_status_pending
                    cell.fill = fill_pending
                elif status == "不符資格":
                    cell.font = font_status_ineligible
                    cell.fill = fill_ineligible
                    
        ws.row_dimensions[curr_row].height = 24
        
    column_min_widths = {
        1: 8,   # 序號
        2: 18,  # 案件編號
        3: 16,  # 大隊/局本部
        4: 16,  # 分隊/科室
        5: 14,  # 申請人
        6: 16,  # 身分證
        7: 14,  # 子女姓名
        8: 14,  # 組別
        9: 14,  # 學期總平均
        10: 12, # 操行
        11: 24, # 附件檢核
        12: 14, # 審核結果
        13: 36, # 判定說明
        14: 16  # 簽章
    }
    
    for col_idx, min_w in column_min_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min_w
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
